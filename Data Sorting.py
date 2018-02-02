import csv,openpyxl,pprint,sys,os,datetime,calendar
import pandas as pd
from openpyxl import workbook, load_workbook

On_Air_Access = 0
On_Air_Duration = datetime.timedelta(seconds=0)
On_Air_Visitors = 0
E24_Access = 0
E24_Duration = datetime.timedelta(seconds=0)
E24_Visitors = 0
News24_Access = 0
News24_Duration = datetime.timedelta(seconds=0)
News24_Visitors = 0
Santa_Access = 0
Santa_Duration = datetime.timedelta(seconds=0)
Santa_Visitors = 0
Web_On_Air_Access = 0
Web_On_Air_Duration = datetime.timedelta(seconds=0)
Web_On_Air_Visitors = 0
Web_E24_Access = 0
Web_E24_Duration = datetime.timedelta(seconds=0)
Web_E24_Visitors = 0
Web_News24_Access = 0
Web_News24_Duration = datetime.timedelta(seconds=0)
Web_News24_Visitors = 0
Web_Santa_Access = 0
Web_Santa_Duration = datetime.timedelta(seconds=0)
Web_Santa_Visitors = 0
IOS_On_Air_Access = 0
IOS_On_Air_Duration = datetime.timedelta(seconds=0)
IOS_On_Air_Visitors = 0
IOS_E24_Access = 0
IOS_E24_Duration = datetime.timedelta(seconds=0)
IOS_E24_Visitors = 0
IOS_News24_Access = 0
IOS_News24_Duration = datetime.timedelta(seconds=0)
IOS_News24_Visitors = 0
IOS_Santa_Access = 0
IOS_Santa_Duration = datetime.timedelta(seconds=0)
IOS_Santa_Visitors = 0
Android_On_Air_Access = 0
Android_On_Air_Duration = datetime.timedelta(seconds=0)
Android_On_Air_Visitors = 0
Android_E24_Access = 0
Android_E24_Duration = datetime.timedelta(seconds=0)
Android_E24_Visitors = 0
Android_News24_Access = 0
Android_News24_Duration = datetime.timedelta(seconds=0)
Android_News24_Visitors = 0
Android_Santa_Access = 0
Android_Santa_Duration = datetime.timedelta(seconds=0)
Android_Santa_Visitors = 0
IR_On_Air_Access = 0
IR_On_Air_Duration = datetime.timedelta(seconds=0)
IR_On_Air_Visitors = 0
IR_E24_Access = 0
IR_E24_Duration = datetime.timedelta(seconds=0)
IR_E24_Visitors = 0
IR_News24_Access = 0
IR_News24_Duration = datetime.timedelta(seconds=0)
IR_News24_Visitors = 0
IR_Santa_Access = 0
IR_Santa_Duration = datetime.timedelta(seconds=0)
IR_Santa_Visitors = 0

On_Air = []
E24 = []
News24 = []
Santa = []
KCRW2016 = []
KCRW2017= []
Website_Air = []
Website_E24 = []
Website_News24 = []
Website_Santa = []
IOS_Air = []
IOS_E24 = []
IOS_News24 = []
IOS_Santa = []
Android_Air = []
Android_E24 = []
Android_News24 = []
Android_Santa = []
IR_Air = []
IR_E24 = []
IR_News24 = []
IR_Santa = []

merginglist = []
lenlist = []
lenlist2017 = []
comblist = []
count = 0
loophelper =0
seperator = 0

Working_Folder = 'APR-DEC2016.xlsx'
# file = list(csv.reader(open('Apr-Dec_2016.csv','r')))
apr_dec2016=[calendar.month_name[x+4]+'2016' for x in range(len(calendar.month_name[4:13]))]
Jan_Present=[calendar.month_name[x+1]+'2017' for x in range(len(calendar.month_name[1:13]))]


# Jan_Present=[calendar.month_name[x+1]+'2017' for x in range(len(calendar.month_name[1:datetime.datetime.now().month]))]

# TODO - Update the name keeping for 2018

for x in range(len(apr_dec2016)):
    try:
        KCRW2016.append(pd.read_csv('KCRW_'+apr_dec2016[x]+'.csv')) # THIS WILL BE THE SAVE SYNTAX!!!!
        merginglist.append(pd.read_csv('KCRW_'+apr_dec2016[x]+'.csv'))
        lenlist.append(len(pd.read_csv('KCRW_'+apr_dec2016[x]+'.csv')))
        comblist.append(len(pd.read_csv('KCRW_'+apr_dec2016[x]+'.csv')))
    except FileNotFoundError:
        print('KCRW_'+apr_dec2016[x]+'.csv',' file not found!, It must not have been downloaded yet, or was saved improperly')

for x in range(len(Jan_Present)):
    try:
        KCRW2017.append(pd.read_csv('C:\\Users\\Uchenna\\Documents\\Python\\KCRW Source Files\\StreamGuys'+'\\KCRW_'+Jan_Present[x]+'.csv')) # THIS WILL BE THE SAVE SYNTAX!!!!
        merginglist.append(pd.read_csv('C:\\Users\\Uchenna\\Documents\\Python\\KCRW Source Files\\StreamGuys'+'\\KCRW_'+Jan_Present[x]+'.csv'))
        lenlist2017.append(len(pd.read_csv('C:\\Users\\Uchenna\\Documents\\Python\\KCRW Source Files\\StreamGuys'+'\\KCRW_'+Jan_Present[x]+'.csv')))
        comblist.append(len(pd.read_csv('C:\\Users\\Uchenna\\Documents\\Python\\KCRW Source Files\\StreamGuys'+'\\KCRW_'+Jan_Present[x]+'.csv')))
    except FileNotFoundError:
        print('KCRW_'+Jan_Present[x]+'.csv',' file not found!, It must not have been downloaded yet, or was saved improperly')

merge = pd.concat(merginglist)
merge.to_csv('BetterTogether.csv')

file = list(csv.reader(open('BetterTogether.csv','r')))

wb= openpyxl.load_workbook(Working_Folder)
if 'Sheet1' in wb.sheetnames:
    wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
        
if 'StreamGuys_Replica' in wb.sheetnames:
    sheet = wb.get_sheet_by_name('StreamGuys_Replica')
else:
    sheet = wb.create_sheet('StreamGuys_Replica')


firstpass = [str(file[x][1]).strip("['/").split('\\') for x in range(len(file))]
slashless = [str(firstpass[x]).strip("['/").strip("']").split('\\') for x in range(len(firstpass))]


suffixes = ['B', 'KB', 'MB', 'GB', 'TB', 'PB']
def humansize(nbytes):
    i = 0
    while nbytes >= 1024 and i < len(suffixes)-1:
        nbytes /= 1024.
        i += 1
    f = ('%.2f' % nbytes).rstrip('0').rstrip('.')
    return '%s %s' % (f, suffixes[i])

Titles = [slashless[0][x].strip('t') for x in range (len(slashless[x]))]
Page  = [slashless[x+1][0].replace('?(parameters)','').replace('192k_mp3','Website').replace('128k_aac','IOS_app').replace('128k_mp3','_Android_App').replace('_internet_radio','IR_Marker') for x in range(len(slashless)-1)]
Accesses = [slashless[x+1][1].strip('t') for x in range (len(slashless)-1)]
Avg_DurationSec = [datetime.timedelta(seconds=int(slashless[x+1][2].strip('t'))) for x in range (len(slashless)-1)]
DurationSec = [datetime.timedelta(seconds=int(slashless[x+1][3].strip('t'))) for x in range (len(slashless)-1)]
SizeBytes = [humansize(int(slashless[x+1][4].strip('t'))) for x in range (len(slashless)-1)]
Visitors = [slashless[x+1][5].strip('t') for x in range (len(slashless)-1)]

BigList = [Page,Accesses,Avg_DurationSec,DurationSec,SizeBytes,Visitors]

count+=1

for x in range(len(Titles)):
    sheet.cell(row=1,column=x+1).value=Titles[x]
  
for y in range(len(lenlist)):
    # sheet.cell(row=lenlist[y],column =7).value=apr_dec2016[y] 
    for x in range (loophelper,(lenlist[y]+loophelper)):     
        sheet.cell(row=x+2,column=1).value=BigList[0][x]    # Pages
        sheet.cell(row=x+2,column=2).value=BigList[1][x]    # Accesses
        sheet.cell(row=x+2,column=3).value=BigList[2][x]    # Avg_DurationSec
        sheet.cell(row=x+2,column=4).value=BigList[3][x]    # DurationSec
        sheet.cell(row=x+2,column=5).value=BigList[4][x]    # SizeBytes
        sheet.cell(row=x+2,column=6).value=BigList[5][x]    # Visitors
        sheet.cell(row=x+2,column=7).value=apr_dec2016[y].replace('2016',' 2016')
        count+=1
        if 'December2016' in apr_dec2016[y]:
            for cy in range(len(lenlist2017)):
                for cyl in range(loophelper,(lenlist[y]+loophelper)):
                    sheet.cell(row=cyl+2,column=1).value=BigList[0][cyl]    # Pages
                    sheet.cell(row=cyl+2,column=2).value=BigList[1][cyl]    # Accesses
                    sheet.cell(row=cyl+2,column=3).value=BigList[2][cyl]    # Avg_DurationSec
                    sheet.cell(row=cyl+2,column=4).value=BigList[3][cyl]    # DurationSec
                    sheet.cell(row=cyl+2,column=5).value=BigList[4][cyl]    # SizeBytes
                    sheet.cell(row=cyl+2,column=6).value=BigList[5][cyl]    # Visitors
                    sheet.cell(row=cyl+2,column=7).value=Jan_Present[cy].replace('2017',' 2017')
    loophelper=count-1
                
         

"""
if sheet.max_row-count == 0:
    pass
else:
    for x in range((count+7),sheet.max_row):
         sheet.cell(row=x+1,column=7).value=''

"""
# The Data should resemble the way it looks on the site at this point.


loophelper = 0
count = 0
for c in range(len(lenlist+lenlist2017)):
    for x in range(loophelper,loophelper+comblist[c]):
        if 'on_air' in BigList[0][x].lower():
            On_Air_Access += int(BigList[1][x])
            On_Air_Duration +=BigList[3][x] 
            On_Air_Visitors += int(BigList[5][x])
        if 'e24' in BigList[0][x].lower():
            E24_Access += int(BigList[1][x])
            E24_Duration += BigList[3][x]
            E24_Visitors += int(BigList[5][x])
        if 'news' in BigList[0][x].lower():
            News24_Access += int(BigList[1][x])
            News24_Duration += BigList[3][x]
            News24_Visitors += int(BigList[5][x])
        if 'santa' in BigList[0][x].lower():
            Santa_Access += int(BigList[1][x])
            Santa_Duration += BigList[3][x]
            Santa_Visitors += int(BigList[5][x])
    # Website Table data sorting
        if 'website_on_air' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                Web_On_Air_Access += int(BigList[1][x])
                Web_On_Air_Duration += BigList[3][x]
                Web_On_Air_Visitors += int(BigList[5][x])
        if 'website_e24' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                
                Web_E24_Access += int(BigList[1][x])
                Web_E24_Duration += BigList[3][x]
                Web_E24_Visitors += int(BigList[5][x])
        if 'website_news' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                Web_News24_Access += int(BigList[1][x])
                Web_News24_Duration += BigList[3][x]
                Web_News24_Visitors += int(BigList[5][x])
        if 'website_santa' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                Web_Santa_Access += int(BigList[1][x])
                Web_Santa_Duration += BigList[3][x]
                Web_Santa_Visitors += int(BigList[5][x])
    # IOS Table data sorting
        if 'ios_app_on_air' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                IOS_On_Air_Access += int(BigList[1][x])
                IOS_On_Air_Duration += BigList[3][x]
                IOS_On_Air_Visitors += int(BigList[5][x])
        if 'ios_app_e24' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                IOS_E24_Access += int(BigList[1][x])
                IOS_E24_Duration += BigList[3][x]
                IOS_E24_Visitors += int(BigList[5][x])
        if 'ios_app_news' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                IOS_News24_Access += int(BigList[1][x])
                IOS_News24_Duration += BigList[3][x]
                IOS_News24_Visitors += int(BigList[5][x])
        if 'ios_app_santa' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                IOS_Santa_Access += int(BigList[1][x])
                IOS_Santa_Duration += BigList[3][x]
                IOS_Santa_Visitors += int(BigList[5][x])
    # Android Table data sorting
        if 'android_app_on_air' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                Android_On_Air_Access += int(BigList[1][x])
                Android_On_Air_Duration += BigList[3][x]
                Android_On_Air_Visitors += int(BigList[5][x])
        if 'android_app_e24' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                Android_E24_Access += int(BigList[1][x])
                Android_E24_Duration += BigList[3][x]
                Android_E24_Visitors += int(BigList[5][x])
        if 'android_app_news' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                Android_News24_Access += int(BigList[1][x])
                Android_News24_Duration += BigList[3][x]
                Android_News24_Visitors += int(BigList[5][x])
        if 'android_app_santa' in BigList[0][x].lower():
            if 'IR_Marker' in BigList[0][x]:
                pass
            else:
                Android_Santa_Access += int(BigList[1][x])
                Android_Santa_Duration = BigList[3][x]
                Android_Santa_Visitors += int(BigList[5][x])
    # Internet Radio Table data sorting
        if 'on_airIR_Marker' in BigList[0][x]:
            IR_On_Air_Access += int(BigList[1][x])
            IR_On_Air_Duration += BigList[3][x]
            IR_On_Air_Visitors += int(BigList[5][x])
        if 'e24IR_Marker' in BigList[0][x]:
            IR_E24_Access += int(BigList[1][x])
            IR_E24_Duration += BigList[3][x]
            IR_E24_Visitors += int(BigList[5][x])
        if 'newsIR_Marker' in BigList[0][x]:
            IR_News24_Access += int(BigList[1][x])
            IR_News24_Duration += BigList[3][x]
            IR_News24_Visitors += int(BigList[5][x])
        if 'santa_barbaraIR_Marker' in BigList[0][x]:
            IR_Santa_Access += int(BigList[1][x])
            IR_Santa_Duration += BigList[3][x]
            IR_Santa_Visitors += int(BigList[5][x])    
        count+=1
        
    loophelper = count
    On_Air.append([On_Air_Access,On_Air_Duration,On_Air_Visitors])
    E24.append([E24_Access,E24_Duration,E24_Visitors])
    News24.append([News24_Access,News24_Duration,News24_Visitors])
    Santa.append([Santa_Access,Santa_Duration,Santa_Visitors])
    # Website category sorting
    Website_Air.append([Web_On_Air_Access,Web_On_Air_Duration,Web_On_Air_Visitors])
    Website_E24.append([Web_E24_Access,Web_E24_Duration,Web_E24_Visitors])
    Website_News24.append([Web_News24_Access,Web_News24_Duration,Web_News24_Visitors])
    Website_Santa.append([Web_Santa_Access,Web_Santa_Duration,Web_Santa_Visitors])
    # IOS category sorting
    IOS_Air.append([IOS_On_Air_Access,IOS_On_Air_Duration,IOS_On_Air_Visitors])
    IOS_E24.append([IOS_E24_Access,IOS_E24_Duration,IOS_E24_Visitors])
    IOS_News24.append([IOS_News24_Access,IOS_News24_Duration,IOS_News24_Visitors])
    IOS_Santa.append([IOS_Santa_Access,IOS_Santa_Duration,IOS_Santa_Visitors])
    # Android category sorting
    Android_Air.append([Android_On_Air_Access,Android_On_Air_Duration,Android_On_Air_Visitors])
    Android_E24.append([Android_E24_Access,Android_E24_Duration,Android_E24_Visitors])
    Android_News24.append([Android_News24_Access,Android_News24_Duration,Android_News24_Visitors])
    Android_Santa.append([Android_Santa_Access,Android_Santa_Duration,Android_Santa_Visitors])
    # Internet Radio category sorting
    IR_Air.append([IR_On_Air_Access,IR_On_Air_Duration,IR_On_Air_Visitors])
    IR_E24.append([IR_E24_Access,IR_E24_Duration,IR_E24_Visitors])
    IR_News24.append([IR_News24_Access,IR_News24_Duration,IR_News24_Visitors])
    IR_Santa.append([IR_Santa_Access,IR_Santa_Duration,IR_Santa_Visitors])
    
    
    On_Air_Access = 0
    On_Air_Duration = datetime.timedelta(seconds=0)
    On_Air_Visitors = 0
    E24_Access = 0
    E24_Duration = datetime.timedelta(seconds=0)
    E24_Visitors = 0
    News24_Access = 0
    News24_Duration = datetime.timedelta(seconds=0)
    News24_Visitors = 0
    Santa_Access = 0
    Santa_Duration = datetime.timedelta(seconds=0)
    Santa_Visitors = 0
    Web_On_Air_Access = 0
    Web_On_Air_Duration = datetime.timedelta(seconds=0)
    Web_On_Air_Visitors = 0
    Web_E24_Access = 0
    Web_E24_Duration = datetime.timedelta(seconds=0)
    Web_E24_Visitors = 0
    Web_News24_Access = 0
    Web_News24_Duration = datetime.timedelta(seconds=0)
    Web_News24_Visitors = 0
    Web_Santa_Access = 0
    Web_Santa_Duration = datetime.timedelta(seconds=0)
    Web_Santa_Visitors = 0
    IOS_On_Air_Access = 0
    IOS_On_Air_Duration = datetime.timedelta(seconds=0)
    IOS_On_Air_Visitors = 0
    IOS_E24_Access = 0
    IOS_E24_Duration = datetime.timedelta(seconds=0)
    IOS_E24_Visitors = 0
    IOS_News24_Access = 0
    IOS_News24_Duration = datetime.timedelta(seconds=0)
    IOS_News24_Visitors = 0
    IOS_Santa_Access = 0
    IOS_Santa_Duration = datetime.timedelta(seconds=0)
    IOS_Santa_Visitors = 0
    Android_On_Air_Access = 0
    Android_On_Air_Duration = datetime.timedelta(seconds=0)
    Android_On_Air_Visitors = 0
    Android_E24_Access = 0
    Android_E24_Duration = datetime.timedelta(seconds=0)
    Android_E24_Visitors = 0
    Android_News24_Access = 0
    Android_News24_Duration = datetime.timedelta(seconds=0)
    Android_News24_Visitors = 0
    Android_Santa_Access = 0
    Android_Santa_Duration = datetime.timedelta(seconds=0)
    Android_Santa_Visitors = 0
    IR_On_Air_Access = 0
    IR_On_Air_Duration = datetime.timedelta(seconds=0)
    IR_On_Air_Visitors = 0
    IR_E24_Access = 0
    IR_E24_Duration = datetime.timedelta(seconds=0)
    IR_E24_Visitors = 0
    IR_News24_Access = 0
    IR_News24_Duration = datetime.timedelta(seconds=0)
    IR_News24_Visitors = 0
    IR_Santa_Access = 0
    IR_Santa_Duration = datetime.timedelta(seconds=0)
    IR_Santa_Visitors = 0

# At this point all the data is where you want it to be.

### ---- Table creation begins. ---- ###

LiveStreamCategories = ['Live Streams Totals','On-Air Accesses','On-Air Duration (hours)', 'On-Air Visitors','E24 Accesses','E24 Duration','E24 Visitors','News24 Accesses','News24 Duration','News24 Visitors','Santa Barbara Accesses','Santa Barbara Duration','Santa Barbara Visitors']
LiveStreamWEBSITE = ['Live-Streams WEBSITE Totals','On-Air Accesses','On-Air Duration (hours)', 'On-Air Visitors','E24 Accesses','E24 Duration','E24 Visitors','News24 Accesses','News24 Duration','News24 Visitors','Santa Barbara Accesses','Santa Barbara Duration','Santa Barbara Visitors']
LiveStreamIOS = ['Live-Streams IOS Apps','On-Air Accesses','On-Air Duration (hours)', 'On-Air Visitors','E24 Accesses','E24 Duration','E24 Visitors','News24 Accesses','News24 Duration','News24 Visitors','Santa Barbara Accesses','Santa Barbara Duration','Santa Barbara Visitors']
LiveStreamAndroid = ['Live-Streams Android App','On-Air Accesses','On-Air Duration (hours)', 'On-Air Visitors','E24 Accesses','E24 Duration','E24 Visitors','News24 Accesses','News24 Duration','News24 Visitors','Santa Barbara Accesses','Santa Barbara Duration','Santa Barbara Visitors']
LiveStreamIR = ['Live-Streams INTERNET RADIO App','On-Air Accesses','On-Air Duration (hours)', 'On-Air Visitors','E24 Accesses','E24 Duration','E24 Visitors','News24 Accesses','News24 Duration','News24 Visitors','Santa Barbara Accesses','Santa Barbara Duration','Santa Barbara Visitors']

if 'KCRW_StreamGuys' in wb.sheetnames:
    sheet2 = wb.get_sheet_by_name('KCRW_StreamGuys')
else:
    sheet2 = wb.create_sheet(title='KCRW_StreamGuys')
    

# sheet2['A1']='Live Streams Totals'

"""
    Add 15 to each new row.
    12 for the information, and 1 for the column labels
    2 extra rows at the end for easier readability by user.
"""


# ROW LABELS FOR ALL TABLES

for x in range(len(LiveStreamCategories)):
    sheet2.cell(row=x+1,column=1).value=LiveStreamCategories[x] # Live Streams Row Labels
    sheet2.cell(row=x+15,column=1).value = LiveStreamWEBSITE[x] # WEBSITE Row Labels
    sheet2.cell(row=x+30,column=1).value = LiveStreamIOS[x]     # IOS Row Labels
    sheet2.cell(row=x+45,column=1).value = LiveStreamAndroid[x] # Android Row Labels
    sheet2.cell(row=x+60,column=1).value = LiveStreamIR[x]      # Internet Radio Row Labels 

# COLUMN LABELS FOR ALL TABLES

count=0
loophelper=0
for x in range(len(apr_dec2016)):
    sheet2.cell(row=1,column=x+2).value=apr_dec2016[x].replace('2016',' 2016')          # Live Streams Column Labels
    sheet2.cell(row=15,column=x+2).value = apr_dec2016[x].replace('2016',' 2016')       # WEBSITE Column Labels
    sheet2.cell(row=30,column=x+2).value = apr_dec2016[x].replace('2016',' 2016')       # IOS Column Labels
    sheet2.cell(row=45,column=x+2).value = apr_dec2016[x].replace('2016',' 2016')       # Android Column Labels
    sheet2.cell(row=60,column=x+2).value = apr_dec2016[x].replace('2016',' 2016')       # Internet Radio Label
    count+=1
for x in range(count,(len(apr_dec2016)+len(Jan_Present))):
    sheet2.cell(row=1,column=loophelper+count+2).value=Jan_Present[loophelper].replace('2017',' 2017')          # Live Streams Column Labels
    sheet2.cell(row=15,column=loophelper+count+2).value = Jan_Present[loophelper].replace('2017',' 2017')       # WEBSITE Column Labels
    sheet2.cell(row=30,column=loophelper+count+2).value = Jan_Present[loophelper].replace('2017',' 2017')       # IOS Column Labels
    sheet2.cell(row=45,column=loophelper+count+2).value = Jan_Present[loophelper].replace('2017',' 2017')       # Android Column Labels
    sheet2.cell(row=60,column=loophelper+count+2).value = Jan_Present[loophelper].replace('2017',' 2017')       # Internet Radio Label
    loophelper+=1

for c in range(len(comblist)):
    try:
        for x in range(len(On_Air)+1):
            sheet2.cell(row=x+2, column=c+2).value = On_Air[c][x]
            sheet2.cell(row=x+5, column=c+2).value = E24[c][x]
            sheet2.cell(row=x+8, column=c+2).value = News24[c][x]
            sheet2.cell(row=x+11, column=c+2).value = Santa[c][x]
        # Website
            sheet2.cell(row=x+16, column=c+2).value = Website_Air[c][x]
            sheet2.cell(row=x+19, column=c+2).value = Website_E24[c][x]
            sheet2.cell(row=x+22, column=c+2).value = Website_News24[c][x]
            sheet2.cell(row=x+25, column=c+2).value = Website_Santa[c][x]
        # IOS
            sheet2.cell(row=x+31, column=c+2).value = IOS_Air[c][x]
            sheet2.cell(row=x+34, column=c+2).value = IOS_E24[c][x]
            sheet2.cell(row=x+37, column=c+2).value = IOS_News24[c][x]
            sheet2.cell(row=x+40, column=c+2).value = IOS_Santa[c][x]
        # Android
            sheet2.cell(row=x+46, column=c+2).value = Android_Air[c][x]
            sheet2.cell(row=x+49, column=c+2).value = Android_E24[c][x]
            sheet2.cell(row=x+52, column=c+2).value = Android_News24[c][x]
            sheet2.cell(row=x+55, column=c+2).value = Android_Santa[c][x]
        # Internet Radio
            sheet2.cell(row=x+61, column=c+2).value = IR_Air[c][x]
            sheet2.cell(row=x+64, column=c+2).value = IR_E24[c][x]
            sheet2.cell(row=x+67, column=c+2).value = IR_News24[c][x]
            sheet2.cell(row=x+70, column=c+2).value = IR_Santa[c][x]
    except IndexError:
        print('index issue')

wb.save(Working_Folder)
